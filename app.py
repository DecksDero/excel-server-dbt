from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from datetime import datetime
from io import BytesIO
import os

app = Flask(__name__)
CORS(app)

MESES = ['enero','febrero','marzo','abril','mayo','junio','julio',
         'agosto','septiembre','octubre','noviembre','diciembre']

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'plantilla.xlsx')

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

@app.route('/generar-excel', methods=['POST'])
def generar_excel():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        wb = load_workbook(TEMPLATE_PATH)
        ws = wb['PRESUPUESTO']

        # Limpiar filas variables
        for row in range(18, 56):
            ws.cell(row=row, column=2).value = None
            ws.cell(row=row, column=9).value = None
        for row in range(57, 107):
            ws.cell(row=row, column=2).value = None
            ws.cell(row=row, column=9).value = None

        # Fecha
        now = datetime.now()
        fecha_str = f"{now.day} de {MESES[now.month-1]} {now.year}"
        patente = (data.get('patente') or '').upper().strip()

        # Encabezado
        ws['B7'] = f'Presupuesto {patente}'
        ws['D7'] = f' Fecha {fecha_str}'

        # Datos vehículo
        ws['C13'] = data.get('marca') or ''
        ws['F13'] = data.get('modelo') or ''
        if data.get('anio'):
            try: ws['I13'] = int(data['anio'])
            except: pass
        if data.get('km'):
            try: ws['F14'] = int(data['km'])
            except: pass
        ws['C14'] = patente
        if data.get('combustible'):
            try: ws['I14'] = int(data['combustible']) / 4
            except: pass

        # Reparaciones
        trabajos = data.get('trabajos') or []
        for i, t in enumerate(trabajos):
            if i < 37:
                ws.cell(row=18+i, column=2).value = t

        # Repuestos
        repuestos = data.get('repuestos') or []
        for i, r in enumerate(repuestos):
            if i < 48:
                ws.cell(row=57+i, column=2).value = r
                ws.cell(row=57+i, column=9).value = 'GAMA'

        # Observaciones
        ws['B113'] = data.get('observaciones') or ''

        # Guardar en memoria
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Presupuesto_{patente}.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
